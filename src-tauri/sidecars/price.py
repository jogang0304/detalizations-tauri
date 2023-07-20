import os
import sys
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from xlcalculator import ModelCompiler, Evaluator, ExcelError, Blank


def calculate(filename: str, sheetName: str, row: int, col: str):
    compiler = ModelCompiler()
    model = compiler.read_and_parse_archive(filename)
    evaluator = Evaluator(model)
    cellAddress = f"{sheetName}!{col}{str(row)}"
    val = evaluator.model.cells.get(cellAddress)
    if not val:
        val = evaluator.evaluate(cellAddress)
    else:
        val = val.value
    if not ExcelError.is_error(val) and not Blank.is_blank(val):
        return str(val)
    else:
        return None


def getNameAndPrice(filename: str, toFindPrice: str, toFindName: str):
    price: None | float = None
    name: None | str = None
    wb = load_workbook(filename, read_only=True, keep_vba=True, keep_links=True)
    worksheets = wb.sheetnames
    if len(worksheets) > 0:
        sheetName = worksheets[0]
        ws = wb[sheetName]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == toFindPrice:
                    priceCellRow = cell.row + 1
                    priceCellCol = cell.column_letter
                    priceStr = calculate(
                        filename, sheetName, priceCellRow, priceCellCol
                    )
                    if priceStr:
                        price = round(float(priceStr), 2)
                if cell.value == toFindName:
                    nameCellRow = cell.row
                    nameCellCol = get_column_letter(cell.column + 1)
                    name = calculate(filename, sheetName, nameCellRow, nameCellCol)
                    if name:
                        name = name.lower()
    return (name, price)


if __name__ == "__main__":
    ans = "error"
    if len(sys.argv) >= 4:
        filePath = sys.argv[1]
        toFindPrice = sys.argv[2]
        toFindName = sys.argv[3]
        result = getNameAndPrice(filePath, toFindPrice, toFindName)
        if result[0] and result[1]:
            ans = f"{result[0]} {result[1]}"
    print(ans)
